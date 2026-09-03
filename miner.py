import os

# Must be before playwright import
if os.getenv("RAILWAY_ENVIRONMENT"):
    os.environ["PLAYWRIGHT_BROWSERS_PATH"] = "/app/pw-browsers"

from playwright.sync_api import sync_playwright
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
import json
import time
import random
import re
import sys
import io
import glob
import csv
import pandas as pd
from datetime import date, datetime, timedelta
from dateutil.relativedelta import relativedelta
from dotenv import load_dotenv

print("Imports done.", flush=True)

dotenv_path = os.path.join(os.path.dirname(__file__), ".env")
load_dotenv(dotenv_path)

USERNAME = os.getenv("MINER_USER")
PASSWORD = os.getenv("MINER_PASSWORD")
ADMIN_URL = "https://evolvemedspa.zenoti.com/Admin/Admin.aspx"
COOKIES_FILE = os.path.join(os.path.dirname(__file__), "cookies.json")
TOKEN_FILE = os.path.join(os.path.dirname(__file__), "token.json")
LOGIN_MAX_ATTEMPTS = 3
LOGIN_RETRY_BACKOFF = 5
REDOWNLOAD_MAX_ATTEMPTS = 2
REDOWNLOAD_RETRY_BACKOFF = 10
# Retry passes over reports that failed their first download.
RETRY_MAX_ATTEMPTS = 2

# Current Stock exports as CSV. Flip to True to download it as a workbook
# (#export_excel_v2) instead; the Excel path and validate_excel() stay in place.
CURRENT_STOCK_AS_EXCEL = False

if not USERNAME or not PASSWORD:
    raise ValueError("MINER_USER and MINER_PASSWORD must be set in the .env file.")

# yesterday = (date.today() - timedelta(days=1)).strftime("%Y-%m-%d")
# START_DATE = yesterday
# END_DATE = yesterday

today = date.today()
date_to = today - timedelta(days=1)
# Anchor the window on date_to, not today: on the 1st of a month date_to has
# already rolled back into the previous month, so anchoring on today would
# shift the whole range forward by one month (Aug 1 -> Jul 1..Jul 31 instead
# of Jun 1..Jul 31).
first_day_of_period = date_to.replace(day=1)
date_from = first_day_of_period - relativedelta(months=1)

START_DATE = date_from.strftime("%Y-%m-%d")
END_DATE = date_to.strftime("%Y-%m-%d")
BKP_START_DATE = date_to.replace(day=1).strftime("%Y-%m-%d")
# Orders window ends at date_to (T-1) and starts one calendar month before it,
# anchored on date_to rather than today: Aug 19 -> Jul 18..Aug 18,
# Aug 20 -> Jul 19..Aug 19. Both endpoints are inclusive, so the span is 31-32
# days rather than a strict month - deliberate, chosen over Jul 19..Aug 18.
ORDERS_START_DATE = (date_to - relativedelta(months=1)).strftime("%Y-%m-%d")

print("Date From:", date_from.strftime("%m/%d/%Y"))
print("Date To  :", date_to.strftime("%m/%d/%Y"))
print("Orders   :", ORDERS_START_DATE, "to", END_DATE)

IS_LOCAL = os.getenv("RAILWAY_ENVIRONMENT") is None

DRIVE_FOLDER_ID = "1wKLZcbe8p9Qpgl9g9KZ4G6bGk__JowY5"
DONE_FOLDER_ID = "1icNO-KvNyolmdOAL7d4HSacKVMR72nmz"
REPORT_FOLDERS = {
    "Attendance": "1YKoroJ8l_YSlQGCEBvp9vJIm8sFZOzX0",
    "Cost of Goods": "1M6xHpZAKtBlu6ageNr2KhZYxOTX9iExg",
    "Sales-Cash": "1FXYnXXQiwQxVAu8IBQOm5GROddoBOXwp",
    "Appointments": "12jqbWWMgpgioR_23KJKLXSvDignwrcP2",
    "Sales-Accrual": "1TBdw_u-ADwb3m6GH-HY4WOVYblIPBxd-",
    "Employee Sales": "123hD_j54WtSPIAjdACWFEGaO877lztRs",
    "Business KPI": "1GjkgXcKrGFqa8l9iM-rW8u2MeRVCaB_M",
    "Memberships": "172HJzXYy_9_qtlgTSlZUgUJmZmT-7qwH",
    "Current Stock": "174ZiUaKjIjEKJNKe75mZXKAwya0F4GNK",
    "Stock Ledger": "1JwZGmMBu-3ZHb67edqOZ8vsj5u9eMRd9",
    "FBAds": "1rs8hu18v64Xml3ZQ4F1Mr5uytZ6V5ppC",
    "GoogleAds": "15Cxii7nKW4XXhJNPjAUd2Y819GxneYNa",
    "Orders": "1snJsOiH3EtI_c4rdti1VHtP5oez0n1U-",
}
SCOPES = ["https://www.googleapis.com/auth/drive.file"]

# Reverse lookup for log lines: a folder id says nothing in a log, the report
# name says which folder was written to.
FOLDER_NAMES = {folder_id: name for name, folder_id in REPORT_FOLDERS.items()}
FOLDER_NAMES[DONE_FOLDER_ID] = "Done"
FOLDER_NAMES.setdefault(DRIVE_FOLDER_ID, "Reports")


def folder_label(folder_id):
    return FOLDER_NAMES.get(folder_id, folder_id)

GSHEET_ID = "1ebRZa2y25O5wuPTdbIOeAqokc3FEawToSBVx5yUthfQ"
GSHEET_TABS = {
    "FBAds": {"gid": 1784599697, "folder_id": "1rs8hu18v64Xml3ZQ4F1Mr5uytZ6V5ppC"},
    "GoogleAds": {"gid": 1126535667, "folder_id": "15Cxii7nKW4XXhJNPjAUd2Y819GxneYNa"},
}


_drive_service = None
_drive_creds = None


def get_drive_service():
    from google.auth.transport.requests import Request

    global _drive_service, _drive_creds

    # Load creds once from env; keep them in-memory so a refreshed access token
    # persists across calls (can't save back to env var on Railway).
    if _drive_creds is None:
        token_json = os.getenv("GOOGLE_TOKEN_JSON")
        if not token_json:
            print("GOOGLE_TOKEN_JSON not set. Skipping upload.")
            return None
        creds_info = json.loads(token_json)
        _drive_creds = Credentials.from_authorized_user_info(creds_info, SCOPES)

    # Refresh only when the live cached creds are actually expired (~once/hour),
    # not on every report. Rebuild the service after a refresh.
    if _drive_creds.expired and _drive_creds.refresh_token:
        print("Refreshing Google OAuth token...")
        _drive_creds.refresh(Request())
        _drive_service = None
        print("Token refreshed.")

    if _drive_service is None:
        _drive_service = build("drive", "v3", credentials=_drive_creds)

    return _drive_service

def upload_to_drive(filepath, folder_id=DRIVE_FOLDER_ID):
    service = get_drive_service()
    if not service:
        return None

    filename = os.path.basename(filepath)

    # Phase 1 (moving existing files to Done) runs upfront in
    # move_existing_reports_to_done() before any download, so no pre-move here.
    print(f"Uploading new file: {filename}")
    file_metadata = {"name": filename, "parents": [folder_id]}
    # Pick the mimetype from the extension: Current Stock uploads a workbook, so a
    # hardcoded text/csv would make Drive mislabel (and fail to preview) it.
    mimetypes_by_ext = {
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".xls": "application/vnd.ms-excel",
        ".csv": "text/csv",
    }
    ext = os.path.splitext(filename)[1].lower()
    mimetype = mimetypes_by_ext.get(ext, "application/octet-stream")
    media = MediaFileUpload(filepath, mimetype=mimetype, resumable=True)
    uploaded = service.files().create(
        body=file_metadata,
        media_body=media,
        fields="id",
    ).execute()
    time.sleep(3)

    print(f"Uploaded to {folder_label(folder_id)} folder: {filename}")

    if not uploaded or not uploaded.get('id'):
        raise Exception(f"Upload failed: {filename}")

    # Post-upload sweep: the Google API client silently retries files().create()
    # on a lost/timed-out response, which can create a SECOND server-side copy
    # even though this call returned once with no error. Keep the file we just
    # got back; move any other same-name copy to Done (can't delete under the
    # drive.file scope). This is the guard that catches single-run duplicates.
    if folder_id != DONE_FOLDER_ID:
        kept_id = uploaded.get("id")
        copies = service.files().list(
            q=f"name='{filename}' and '{folder_id}' in parents and trashed=false",
            fields="files(id, name)",
        ).execute().get("files", [])
        for c in copies:
            if c["id"] == kept_id:
                continue
            print(f"Duplicate copy detected: moving extra {filename} to Done")
            try:
                service.files().update(
                    fileId=c["id"],
                    addParents=DONE_FOLDER_ID,
                    removeParents=folder_id,
                    fields="id",
                ).execute()
                time.sleep(1)
            except Exception as sweep_err:
                print(f"  Could not move duplicate {filename}: {sweep_err}")

    return uploaded


def list_all_files(service, folder_id):
    """Every file in a folder, following pagination.

    files().list() returns at most pageSize results (Drive v3 defaults to 100),
    so a single call silently misses anything past the first page. Ask for the
    max and follow nextPageToken so callers really do see all files.

    Caveat: under the drive.file scope this only ever returns files created by
    THIS OAuth client. A file uploaded by hand, or by a run using a different
    client_id, is invisible here and therefore unmovable."""
    files = []
    page_token = None
    while True:
        response = service.files().list(
            q=f"'{folder_id}' in parents and trashed=false",
            fields="nextPageToken, files(id, name)",
            pageSize=1000,
            pageToken=page_token,
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        ).execute()
        files.extend(response.get("files", []))
        page_token = response.get("nextPageToken")
        if not page_token:
            return files


def move_existing_reports_to_done():
    for folder_name, folder_id in REPORT_FOLDERS.items():
        service = get_drive_service()
        if not service:
            return

        # A bad/unreachable folder id raises 404 here. This sweep runs pre-login,
        # so letting it propagate kills the whole run before a single report is
        # downloaded - degrade to a warning and keep the other folders going.
        try:
            existing = list_all_files(service, folder_id)
        except Exception as e:
            print(f"  Could not list {folder_name} ({folder_id}): {e}")
            continue
        if not existing:
            continue
        print(f"{folder_name}: {len(existing)} existing file(s) to move")
        moved = 0
        for old_file in existing:
            print(f"Moving {folder_name}/{old_file['name']} to Done folder")
            # Isolate per file: one un-movable file (permissions, a concurrent
            # delete) shouldn't abandon the rest of this folder or the folders
            # after it. This sweep runs pre-login, so raising here would kill
            # the whole run before a single report is downloaded.
            try:
                service.files().update(
                    fileId=old_file["id"],
                    addParents=DONE_FOLDER_ID,
                    removeParents=folder_id,
                    fields="id, parents",
                    supportsAllDrives=True,
                ).execute()
                moved += 1
                time.sleep(2)
            except Exception as e:
                print(f"  Could not move {folder_name}/{old_file['name']}: {e}")

        if moved != len(existing):
            print(f"  {folder_name}: moved {moved}/{len(existing)}, {len(existing) - moved} left behind")

        # Verify rather than trust the update() response: re-list the folder and
        # report anything still visible. A file that survives here was either
        # re-parented back, or update() reported success without taking effect.
        try:
            remaining = list_all_files(service, folder_id)
        except Exception as e:
            print(f"  {folder_name}: could not verify sweep: {e}")
            continue

        if remaining:
            print(
                f"  WARNING {folder_name}: {len(remaining)} file(s) STILL in the folder "
                f"after the sweep: {[f['name'] for f in remaining]}"
            )
        else:
            print(f"  {folder_name}: verified empty.")


def dedupe_report_folders():
    """Guardrail: if a report folder holds more than one file with the same
    name (e.g. a container restart re-ran the script and re-uploaded), keep one
    and move the extras to Done. Move — not delete — because the drive.file
    scope cannot delete files it did not create."""
    service = get_drive_service()
    if not service:
        return

    for folder_name, folder_id in REPORT_FOLDERS.items():
        files = list_all_files(service, folder_id)

        by_name = {}
        for f in files:
            by_name.setdefault(f["name"], []).append(f)

        for name, dupes in by_name.items():
            if len(dupes) < 2:
                continue
            # keep the first, move the rest to Done
            for dup in dupes[1:]:
                print(f"Duplicate in {folder_name}: moving extra {name} to Done folder")
                try:
                    service.files().update(
                        fileId=dup["id"],
                        addParents=DONE_FOLDER_ID,
                        removeParents=folder_id,
                        fields="id",
                    ).execute()
                    time.sleep(2)
                except Exception as e:
                    print(f"  Could not move duplicate {name}: {e}")


def validate_csv(filepath):
    filename = os.path.basename(filepath)

    if not os.path.exists(filepath):
        raise Exception(f"File not found: {filename}")

    filesize = os.path.getsize(filepath)
    if filesize == 0:
        raise Exception(f"File is empty: {filename}")

    filesize_mb = filesize / (1024 * 1024)
    print(f"  File size: {filesize_mb:.2f} MB")

    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        head = f.read(1024)

    if "<html" in head.lower() or "<!doctype" in head.lower():
        raise Exception(f"File is HTML, not CSV (possible error page): {filename}")

    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        try:
            headers = next(reader)
        except StopIteration:
            raise Exception(f"CSV has no headers: {filename}")

        if len(headers) < 2:
            raise Exception(f"CSV has only {len(headers)} column(s), likely corrupt: {filename}")

        row_count = 0
        for row in reader:
            row_count += 1
            if row_count >= 5:
                break

        if row_count == 0:
            raise Exception(f"CSV has headers but no data rows: {filename}")

    print(f"  CSV valid: {len(headers)} columns, {row_count}+ data rows")
    return True


def validate_excel(filepath):
    """Sanity-check a downloaded workbook.

    Zenoti sometimes serves an HTML error page with a spreadsheet filename, so
    check the magic bytes: .xlsx is a zip ("PK"), legacy .xls is OLE2. Row/column
    counts are only checked when openpyxl is installed.
    """
    filename = os.path.basename(filepath)

    if not os.path.exists(filepath):
        raise Exception(f"File not found: {filename}")

    filesize = os.path.getsize(filepath)
    if filesize == 0:
        raise Exception(f"File is empty: {filename}")

    print(f"  File size: {filesize / (1024 * 1024):.2f} MB")

    with open(filepath, "rb") as f:
        magic = f.read(8)

    if magic[:2] == b"PK":
        kind = "xlsx"
    elif magic == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        kind = "xls"
    else:
        head = magic.decode("utf-8", errors="replace").lower()
        if "<html" in head or "<!doc" in head or magic.lstrip()[:1] == b"<":
            raise Exception(f"File is HTML, not a workbook (possible error page): {filename}")
        raise Exception(f"File is not a valid Excel workbook: {filename}")

    if kind == "xls":
        # Legacy format; openpyxl cannot read it. Magic bytes are as far as we go.
        print(f"  Workbook valid: legacy .xls, {filesize} bytes")
        return True

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  Workbook valid: xlsx container OK (install openpyxl for row checks)")
        return True

    wb = load_workbook(filepath, read_only=True)
    try:
        ws = wb.active
        rows = ws.max_row or 0
        cols = ws.max_column or 0
        if cols < 2:
            raise Exception(f"Workbook has only {cols} column(s), likely corrupt: {filename}")
        if rows < 2:
            raise Exception(f"Workbook has headers but no data rows: {filename}")
    finally:
        wb.close()

    print(f"  Workbook valid: {cols} columns, {rows} rows")
    return True


def cleanup_old_csvs():
    script_dir = os.path.dirname(__file__) or "."
    # Current Stock now downloads as a workbook, so sweep those too — otherwise
    # a stale current_stock_*.xlsx sits in the repo root forever.
    for pattern in ("*.csv", "*.xlsx", "*.xls"):
        for f in glob.glob(os.path.join(script_dir, pattern)):
            if date_to.strftime("%Y-%m-%d") not in os.path.basename(f):
                os.remove(f)
                print(f"Cleaned up old report: {f}")


def create_browser_and_context(pw):
    launch_args = {
        "headless": True,
        "args": [
            "--start-maximized",
            "--disable-blink-features=AutomationControlled",
        ],
    }

    if IS_LOCAL:
        launch_args["channel"] = "chrome"
    else:
        launch_args["args"] += [
            "--no-sandbox",
            "--disable-dev-shm-usage",
        ]

    browser = pw.chromium.launch(**launch_args)

    context_args = {"no_viewport": True, "user_agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"}
    if os.path.exists(COOKIES_FILE):
        print(f"Loading saved cookies from {COOKIES_FILE}")
        context_args["storage_state"] = COOKIES_FILE

    context = browser.new_context(**context_args)
    context.add_init_script("""
        Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
        window.chrome = {runtime: {}};
    """)
    return browser, context


def save_cookies(context):
    context.storage_state(path=COOKIES_FILE)
    print(f"Cookies saved to {COOKIES_FILE}")


def needs_login(page):
    # print(f"Checking if login is needed. Current URL: {page.url}")
    page.goto(ADMIN_URL, wait_until="domcontentloaded")
    try:
        page.wait_for_url("**/Admin/**", timeout=10000)
        return False
    except:
        return True


def _login_attempt(page):
    print(f"Current URL before login: {page.url}")

    username_sel = "input#Username, input[name='Username'], input[name='username'], input[type='email']"
    if not page.locator(username_sel).first.is_visible():
        print("Login form not visible, navigating to admin...")
        page.goto(ADMIN_URL, wait_until="networkidle")

    try:
        page.wait_for_selector(username_sel, state="visible", timeout=30000)
    except Exception as e:
        print(f"Login form not found. URL: {page.url}")
        print(f"Page title: {page.title()}")
        print(f"Page content preview: {page.content()[:1000]}")
        raise e
    print(f"Login page loaded. URL: {page.url}")

    page.locator(username_sel).first.click()
    page.locator(username_sel).first.fill("")
    page.locator(username_sel).first.press_sequentially(USERNAME, delay=50)
    time.sleep(random.uniform(0.5, 1.5))
    print("Username entered.")

    page.locator('#Password').click()
    page.locator('#Password').fill("")
    page.locator('#Password').press_sequentially(PASSWORD, delay=50)
    time.sleep(random.uniform(0.5, 1.5))
    print("Password entered.")
    time.sleep(2)

    login_button = page.locator('#btnLogin')
    print("Waiting for login button...")
    try:
        login_button.click(timeout=10000)
    except:
        print("Button disabled (captcha pending). Forcing submit via JS...")
        page.evaluate("document.getElementById('btnLogin').removeAttribute('disabled')")
        page.evaluate("document.getElementById('btnLogin').click()")
    print("Login button clicked.")
    time.sleep(random.uniform(2.0, 3.0))

    page.wait_for_url("**/Admin/**", timeout=30000)
    print("Login successful!")


def do_login(page):
    """Log in, retrying up to LOGIN_MAX_ATTEMPTS times.

    Login is a single point of failure for the whole run: all three call sites
    abandon their remaining work if it raises. Transient causes (slow login
    form, captcha interstitial, a dropped nav) usually clear on a second try."""
    last_error = None

    for attempt in range(1, LOGIN_MAX_ATTEMPTS + 1):
        print(f"Login attempt {attempt}/{LOGIN_MAX_ATTEMPTS}...")
        try:
            _login_attempt(page)
            return
        except Exception as e:
            last_error = e
            print(f"Login attempt {attempt}/{LOGIN_MAX_ATTEMPTS} failed: {e}")

        if attempt == LOGIN_MAX_ATTEMPTS:
            break

        backoff = LOGIN_RETRY_BACKOFF * attempt
        print(f"Retrying login in {backoff}s...")
        time.sleep(backoff)

        # Reload the login page so the next attempt starts from a clean form
        # rather than whatever half-submitted state the failure left behind.
        try:
            page.goto(ADMIN_URL, wait_until="networkidle", timeout=60000)
        except Exception as nav_err:
            print(f"  Could not reload login page: {nav_err}")
            continue

        # The submit may have actually gone through and only the post-login
        # wait timed out — same check needs_login() uses.
        if "/Admin/" in page.url:
            print("Already authenticated after reload. Login complete.")
            return

    raise Exception(
        f"Login failed after {LOGIN_MAX_ATTEMPTS} attempts: {last_error}"
    ) from last_error


def wait_for_dashboard(page):
    # Login lands on either shell depending on the session: the Reports
    # dashboard (#menuLinkreports) or the admin rail (div.menuNavBtn). Accept
    # any nav element - every caller only needs "the shell is up", and
    # download_report() navigates to the page it wants itself.
    try:
        page.locator('#menuLinkreports, div.menuNavBtn, #usernameBtn').first.wait_for(state='visible', timeout=60000)
        print("Dashboard loaded.")
    except Exception as e:
        print(f"Error: Dashboard menu not found. URL: {page.url}")
        raise e


def apply_appointments_filters(report_page):
    print("  Applying Appointments filters...")
    report_page.evaluate("""
        (function() {
            // All multi-selects → selectAll (Centers, Appointment Status, Appointment Source)
            $('select[multiple]').each(function() {
                $(this).multiselect('selectAll', false);
            });
            // Date Type (single select) → Appointment Date
            $('select:not([multiple])').each(function() {
                var opts = Array.from(this.options);
                var match = opts.find(function(o) { return o.text.trim().indexOf('Appointment Date') !== -1; });
                if (match) {
                    $(this).multiselect('select', match.value);
                }
            });
        })();
    """)
    time.sleep(2)
    print("  Appointments filters applied.")


def apply_attendance_filters(report_page):
    print("  Applying Attendance filters...")
    report_page.evaluate("""
        (function() {
            // Zenoti dropdowns: Centers + Employee Jobs → All
            ['elm_centers', 'elm_employee_jobs'].forEach(function(id) {
                var cb = document.getElementById(id + '-zenoti-dropdown-options-all');
                if (cb && !cb.checked) cb.click();
            });

            // All multi-selects → selectAll
            $('select[multiple]').each(function() {
                $(this).multiselect('selectAll', false);
            });

            // Override: Schedule Status → Working only
            $('select').each(function() {
                var opts = Array.from(this.options);
                if (opts.some(function(o) { return o.value === '6a9d2c87-d452-471f-ba33-90af26ae4edb'; })) {
                    $(this).multiselect('deselectAll', false);
                    $(this).multiselect('select', ['6a9d2c87-d452-471f-ba33-90af26ae4edb']);
                }
            });

            // Single selects: View By → Date, Check-in/Checkout Status → All
            $('select:not([multiple])').each(function() {
                var opts = Array.from(this.options);
                var texts = opts.map(function(o) { return o.text.trim(); });
                if (texts.indexOf('Date') !== -1 && texts.indexOf('Check-in') !== -1) {
                    $(this).multiselect('select', '1');
                } else if (texts.some(function(t) { return t.indexOf('Missed check-ins') !== -1; })) {
                    $(this).multiselect('select', '0');
                }
            });
        })();
    """)
    time.sleep(2)
    print("  Attendance filters applied.")


def apply_cost_of_goods_filters(report_page):
    print("  Applying Cost of Goods filters...")
    report_page.evaluate("""
        (function() {
            // All multi-selects → selectAll (Centers, Product Type, Consumption Type, Brand, Category, Sub Category, Business Unit)
            $('select[multiple]').each(function() {
                $(this).multiselect('selectAll', false);
            });
            // Stock Costing Method (single select) → Perpetual Average Cost
            $('select:not([multiple])').each(function() {
                var opts = Array.from(this.options);
                if (opts.some(function(o) { return o.text.indexOf('Perpetual') !== -1; })) {
                    $(this).multiselect('select', '1');
                }
            });
        })();
    """)
    time.sleep(2)
    print("  Cost of Goods filters applied.")


def apply_sales_accrual_filters(report_page):
    print("  Applying Sales-Accrual filters...")
    report_page.evaluate("""
        (function() {
            // Zenoti dropdown: Centers → All
            var cb = document.getElementById('elm_centers-zenoti-dropdown-options-all');
            if (cb && !cb.checked) cb.click();

            // All multi-selects → selectAll (Category, Sub Category, Business Unit, Payment Type, Sale Type, Invoice Status)
            $('select[multiple]').each(function() {
                $(this).multiselect('selectAll', false);
            });

            // Override: Item Type → Service + Product only
            var $itemType = $('#elm_item_type');
            if ($itemType.length) {
                $itemType.multiselect('deselectAll', false);
                $itemType.multiselect('select', ['Service', 'Product']);
            }
        })();
    """)
    time.sleep(1)
    selected = report_page.evaluate(
        "Array.from(document.querySelectorAll('#elm_item_type option')).filter(function(o){return o.selected}).map(function(o){return o.value})"
    )
    print(f"  Item Type selected: {selected}")
    time.sleep(2)
    print("  Sales-Accrual filters applied.")


def apply_employee_sales_filters(report_page):
    print("  Applying Employee Sales filters...")
    # Targeted per-id selection, NOT a blanket $('select[multiple]') sweep:
    # the Employee filter is also a multiselect and must stay empty.
    report_page.evaluate("""
        (function() {
            ['elm_allowed_centers', 'elm_payment_type', 'elm_sale_type', 'elm_invoice_status'].forEach(function(id) {
                var $sel = $('#' + id);
                if ($sel.length) {
                    $sel.multiselect('selectAll', false);
                    $sel.multiselect('updateButtonText');
                }
            });

            // Item Type → Service + Product only
            var $itemType = $('#elm_item_type');
            if ($itemType.length) {
                $itemType.multiselect('deselectAll', false);
                $itemType.multiselect('select', ['Service', 'Product']);
                $itemType.multiselect('updateButtonText');
            }
        })();
    """)
    time.sleep(2)

    selected = report_page.evaluate("""
        (function() {
            var out = {};
            ['elm_allowed_centers', 'elm_item_type', 'elm_payment_type', 'elm_sale_type', 'elm_invoice_status'].forEach(function(id) {
                var el = document.getElementById(id);
                out[id] = el ? Array.from(el.options).filter(function(o){return o.selected}).length : null;
            });
            return out;
        })();
    """)
    print(f"  Selected counts: {selected}")
    print("  Employee filter left empty.")
    print("  Employee Sales filters applied.")


def apply_sales_cash_filters(report_page):
    print("  Applying Sales-Cash filters...")
    report_page.evaluate("""
        (function() {
            // Zenoti dropdown: Centers → All
            var cb = document.getElementById('elm_centers-zenoti-dropdown-options-all');
            if (cb && !cb.checked) cb.click();

            // Level of Detail (single select) → Item
            var $lod = $('#elm_level_of_detail');
            if ($lod.length) {
                $lod.multiselect('select', '1');
            }

            // All multi-selects → selectAll (Item Type, Category, Sub Category, Business Unit, Sale Type, Invoice Status)
            $('select[multiple]').each(function() {
                $(this).multiselect('selectAll', false);
            });

            // Override: Payment Type → Cash, Card, Check, Custom-Financial, CustomNon-Financial only
            $('select[multiple]').each(function() {
                var values = Array.from(this.options).map(function(o) { return o.value; });
                if (values.indexOf('16') !== -1 && values.indexOf('10') !== -1) {
                    $(this).multiselect('deselectAll', false);
                    $(this).multiselect('select', ['0', '1', '2', '3', '4']);
                }
            });
        })();
    """)
    time.sleep(2)
    print("  Sales-Cash filters applied.")


def apply_business_kpi_filters(report_page):
    print("  Applying Business KPI filters...")
    report_page.evaluate("""
        (function() {
            // Centers → All
            var cb = document.getElementById('elm_centers-zenoti-dropdown-options-all');
            if (cb && !cb.checked) cb.click();

            // Invoice Status → select All
            $('select[multiple]').each(function() {
                $(this).multiselect('selectAll', false);
            });

            // Uncheck 'Show Sales Including Tax'
            var taxCb = document.getElementById('elm_include_tax');
            if (taxCb && taxCb.checked) taxCb.click();
        })();
    """)
    time.sleep(2)
    print("  Business KPI filters applied.")


def apply_memberships_filters(report_page):
    print("  Applying Memberships filters...")
    report_page.evaluate("""
        (function() {
            // All multi-selects → selectAll (Sale Centers, Membership Type, Membership Stats)
            $('select[multiple]').each(function() {
                $(this).multiselect('selectAll', false);
            });

            // Liability Type → By Sale (value "1")
            $('select:not([multiple])').each(function() {
                var opts = Array.from(this.options);
                if (opts.some(function(o) { return o.text.indexOf('By Sale') !== -1 && o.text.indexOf('By Value') === -1; })) {
                    $(this).multiselect('select', '1');
                }
            });

            // Date Type → Sale Date (value "2")
            $('select:not([multiple])').each(function() {
                var opts = Array.from(this.options);
                if (opts.some(function(o) { return o.text.indexOf('Sale Date') !== -1; }) &&
                    opts.some(function(o) { return o.text.indexOf('Balance As On Date') !== -1; })) {
                    $(this).multiselect('select', '2');
                }
            });

            // Status Type → Membership
            var $statusType = $('#elm_status_type');
            if ($statusType.length) {
                $statusType.multiselect('select', 'Membership');
            }
        })();
    """)
    time.sleep(2)
    print("  Memberships filters applied.")


def apply_current_stock_filters(report_page):
    print("  Applying Current Stock filters...")
    report_page.evaluate("""
        (function() {
            // All multi-selects → selectAll (Centers, Category, Sub Category,
            // Vendor, Brand, Business Unit)
            $('select[multiple]').each(function() {
                $(this).multiselect('selectAll', false);
            });

            // Single selects (radios)
            $('select:not([multiple])').each(function() {
                var opts = Array.from(this.options);
                var texts = opts.map(function(o) { return o.text.trim(); });

                // Product Type → All (value "3")
                if (texts.indexOf('Retail') !== -1 && texts.indexOf('Consumable') !== -1) {
                    $(this).multiselect('select', '3');
                }
                // On-Hand Qty → All (value "0")
                else if (texts.indexOf('Greater than 0') !== -1 && texts.indexOf('Less than 0') !== -1) {
                    $(this).multiselect('select', '0');
                }
            });
        })();
    """)
    time.sleep(2)
    print("  Current Stock filters applied.")


def apply_stock_ledger_filters(report_page):
    print("  Applying Stock Ledger filters...")
    report_page.evaluate("""
        (function() {
            // Zenoti dropdown: Centers → All
            var cb = document.getElementById('elm_centers-zenoti-dropdown-options-all');
            if (cb && !cb.checked) cb.click();

            // All multi-selects → selectAll (Category, Sub Category, Product Type,
            // Vendor, Brand, Transaction Type, Business Unit)
            $('select[multiple]').each(function() {
                $(this).multiselect('selectAll', false);
            });
            // Stock Costing Method (single select) → Perpetual Average Cost
            $('select:not([multiple])').each(function() {
                var opts = Array.from(this.options);
                if (opts.some(function(o) { return o.text.indexOf('Perpetual') !== -1; })) {
                    $(this).multiselect('select', '1');
                }
            });
        })();
    """)
    time.sleep(2)
    print("  Stock Ledger filters applied.")


def download_gsheet_reports():
    script_dir = os.path.dirname(__file__) or "."
    succeeded = []
    failed = []

    for tab_name, info in GSHEET_TABS.items():
        try:
            url = f"https://docs.google.com/spreadsheets/d/{GSHEET_ID}/export?format=csv&gid={info['gid']}"
            print(f"\nFetching Google Sheet: {tab_name} (gid={info['gid']})")
            df = pd.read_csv(url)
            print(f"  Rows: {len(df)}, Columns: {len(df.columns)}")

            filename = os.path.join(script_dir, f"{tab_name}_{END_DATE}.csv")
            df.to_csv(filename, index=False)

            validate_csv(filename)
            upload_to_drive(filename, info["folder_id"])
            os.remove(filename)
            succeeded.append(tab_name)
        except Exception as e:
            print(f"FAILED Google Sheet: {tab_name} — {e}")
            failed.append((tab_name, str(e)))

    return succeeded, failed


def validate_report_folders(succeeded_reports):
    service = get_drive_service()
    if not service:
        return []

    missing = []
    for report in succeeded_reports:
        folder_id = REPORT_FOLDERS.get(report)
        if not folder_id:
            continue

        if report == "Business KPI":
            # report_window() sends (BKP_START_DATE, END_DATE) and download_report()
            # names the file from both endpoints, so the range form is what lands.
            expected = [f"business_kpi_{BKP_START_DATE}_to_{END_DATE}.csv"]
        elif report == "Current Stock":
            if CURRENT_STOCK_AS_EXCEL:
                # Extension follows whatever the server served, so accept both.
                expected = [
                    f"current_stock_{END_DATE}.xlsx",
                    f"current_stock_{END_DATE}.xls",
                ]
            else:
                expected = [f"current_stock_{END_DATE}.csv"]
        elif report == "Orders":
            expected = [f"orders_{END_DATE}.csv"]
        else:
            safe_name = report.replace(" ", "_").lower()
            expected = [f"{safe_name}_{START_DATE}_to_{END_DATE}.csv"]

        results = service.files().list(
            q=f"'{folder_id}' in parents and trashed=false",
            fields="files(id, name)",
            pageSize=1000,
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        ).execute()
        files = results.get("files", [])
        names = [f["name"] for f in files]

        found = next((e for e in expected if e in names), None)
        if found:
            print(f"  OK {report}: found '{found}'")
        else:
            print(f"  MISSING {report}: expected '{' or '.join(expected)}' — found: {names}")
            missing.append(report)

    if missing:
        print(f"Missing reports in Drive: {missing}")
    else:
        print("All report folders validated successfully.")
    return missing


# --- Inventory > Orders (PODetailsV2) ----------------------------------------
# Orders is the one report that does not live on ReportsDashboard.aspx: it is an
# ASP.NET listing page with an ag-Grid, entered by clicking the Inventory nav
# (menuNavigationClick). A direct goto to
# https://evolvemedspa.zenoti.com/ListingPages/PODetailsV2.aspx does not work.
# Its filters are a zfc dropdown + two jQuery UI datepickers and its export is a
# __doPostBack link, so none of the report code above applies to it.

# The RadMenu center picker must sit at org level, not on one center: a single
# center scopes the grid - and the export - to that center only.
EXPECTED_CENTER = "Evolve Med Spa"

# Consecutive empty grid readings needed before calling a window "no data":
# a mid-refresh grid also renders zero rows for a moment.
GRID_EMPTY_CONFIRMATIONS = 3

# Returned by download_orders() instead of a filename when the window genuinely
# holds no orders. Exporting an empty grid yields a headers-only file, which
# validates and looks current while reporting zero orders downstream.
ORDERS_NO_DATA = object()

INVENTORY_NAV_SELS = (
    "div.menuNavBtn[menu-item-key='inventory']",
    "div.menuNavBtn[menu-module='inventory']",
    "div.menuNavBtn[menu-url*='InventoryHome.aspx']",
)
ORDERS_LINK_SELS = (
    "#leftAdminFlyPanel a[href='/ListingPages/PODetailsV2.aspx']",
    "#leftAdminFlyPanel a[href*='PODetailsV2.aspx']",
    "div.menuDashCls a[href*='PODetailsV2.aspx']",
    "a[href*='PODetailsV2.aspx']",
)


def _mdy(date_str):
    """YYYY-MM-DD -> M/d/yyyy, the Orders datepicker placeholder (unpadded)."""
    d = datetime.strptime(date_str, "%Y-%m-%d").date()
    return f"{d.month}/{d.day}/{d.year}"


def find_in_frames(page, selectors, timeout=30000):
    """Return (frame, selector) for the first visible match in any frame."""
    deadline = time.time() + timeout / 1000.0
    while True:
        for frame in page.frames:
            for sel in selectors:
                try:
                    if frame.locator(sel).first.is_visible(timeout=500):
                        return frame, sel
                except Exception:
                    continue
        if time.time() >= deadline:
            return None, None
        time.sleep(1)


def settle(page, timeout=120000):
    try:
        page.wait_for_load_state("load", timeout=timeout)
    except Exception as e:
        print(f"  load state not reached: {e}")
    try:
        page.wait_for_load_state("networkidle", timeout=timeout)
    except Exception as e:
        print(f"  networkidle not reached: {e}")


def _nav_to_orders_once(page):
    """Inventory nav button -> fly-out panel -> Orders link."""
    # Always reload the shell, even when already under /Admin/. Orders runs last,
    # so the page is normally still ReportsDashboard.aspx with its #dialog-reports
    # modal open - and that modal intercepts pointer events, so the Inventory nav
    # click times out on an element Playwright reports as visible and enabled.
    if page.url != ADMIN_URL:
        print(f"  Loading a clean admin shell (was {page.url}).")
        page.goto(ADMIN_URL, wait_until="domcontentloaded", timeout=120000)
        settle(page)

    frame, sel = find_in_frames(page, INVENTORY_NAV_SELS, timeout=60000)
    if not frame:
        raise Exception(f"Inventory nav button not found. URL: {page.url}")
    frame.locator(sel).first.click(timeout=15000)
    print(f"  Clicked Inventory nav ({sel}).")
    time.sleep(2)
    # The click may also navigate to InventoryHome.aspx; let that finish.
    settle(page)

    # The fly-out is informational here - the Orders link is what must be clickable.
    panel_frame, _ = find_in_frames(page, ("#leftAdminFlyPanel",), timeout=15000)
    print(f"  Inventory fly-out panel: {'visible' if panel_frame else 'not detected'}")

    frame, sel = find_in_frames(page, ORDERS_LINK_SELS, timeout=60000)
    if not frame:
        raise Exception(f"Orders link not found after clicking Inventory. URL: {page.url}")
    frame.locator(sel).first.click(timeout=15000)
    print(f"  Clicked Orders link ({sel}).")
    time.sleep(2)
    settle(page)


def open_orders_page(page, attempts=2):
    """Reach the Orders listing through the Inventory nav panel."""
    print("Opening Inventory > Orders...")
    last_error = None

    for attempt in range(1, attempts + 1):
        try:
            _nav_to_orders_once(page)
        except Exception as e:
            last_error = e
            print(f"  Nav attempt {attempt}/{attempts} failed: {e}")

        if "PODetailsV2" in page.url:
            break

        if page.url and "PODetailsV2" not in page.url:
            print(f"  Still not on Orders after attempt {attempt}/{attempts}. URL: {page.url}")
        if attempt < attempts:
            time.sleep(5)

    if "PODetailsV2" not in page.url:
        raise Exception(
            f"Could not reach the Orders page via the Inventory nav. URL: {page.url}"
            + (f" Last error: {last_error}" if last_error else "")
        )

    # The date/export controls are rendered late; wait on them, not just the URL.
    frame, _ = find_in_frames(page, ("#ddlTimePeriod",), timeout=120000)
    if not frame:
        raise Exception(f"Orders page loaded but #ddlTimePeriod never appeared. URL: {page.url}")
    time.sleep(2)
    print(f"Orders page fully loaded. URL: {page.url}")
    return frame


def _norm_center(text):
    return "".join(ch for ch in (text or "").lower() if ch.isalnum())


def ensure_center_scope(page):
    """Guard the center picker: the RadMenu root must read the org name.

    A single center selected in that menu silently scopes the grid - and the
    export - to that one center, so a wrong scope must fail the run rather than
    produce a whole-org filename holding one center's orders.
    """
    print(f"Checking center scope (expecting '{EXPECTED_CENTER}')...")
    if "/Admin/" not in page.url:
        print(f"  Not on the admin shell ({page.url}); loading it first.")
        page.goto(ADMIN_URL, wait_until="domcontentloaded", timeout=120000)
        settle(page)
    root_sels = ("a.rmRootLink span.rmText", "a.rmRootLink", ".rmRootLink .rmText")
    frame, sel = find_in_frames(page, root_sels, timeout=60000)
    if not frame:
        raise Exception("Center picker (a.rmRootLink) not found; cannot confirm the center scope")

    current = (frame.locator(sel).first.text_content() or "").strip()
    if _norm_center(current) == _norm_center(EXPECTED_CENTER):
        print(f"  Center scope OK: '{current}'")
        return

    print(f"  Center scope is '{current}', not '{EXPECTED_CENTER}'. Trying to switch...")
    # The item is <a class="rmLink" href="#"> with no onclick: the handler lives
    # on the Telerik RadMenu client object. So el.click() hits nothing, and a
    # real click cannot land either - this menu does not open on hover, so the
    # item stays hidden. Drive the widget's own API, and fall back to a full
    # bubbling mouse sequence if the client object is unreachable.
    switch = frame.evaluate(
        """(wanted) => {
            const norm = s => (s || '').toLowerCase().replace(/[^a-z0-9]/g, '');
            const target = norm(wanted);
            const result = {menus: 0, found: false, method: null, error: null, options: []};

            // get_items() returns a Telerik RadMenuItemCollection, not an array:
            // it exposes get_count()/getItem(i) and is not iterable.
            const toArray = coll => {
                if (!coll) return [];
                if (Array.isArray(coll)) return coll;
                if (typeof coll.toArray === 'function') return coll.toArray();
                if (typeof coll.get_count === 'function' && typeof coll.getItem === 'function') {
                    const out = [];
                    for (let i = 0; i < coll.get_count(); i++) out.push(coll.getItem(i));
                    return out;
                }
                return [];
            };

            const walk = (coll, acc) => {
                for (const it of toArray(coll)) {
                    acc.push(it);
                    if (typeof it.get_items === 'function') walk(it.get_items(), acc);
                }
                return acc;
            };

            if (typeof $find === 'function') {
                const roots = Array.from(document.querySelectorAll('[id]'))
                    .filter(el => String(el.className || '').indexOf('RadMenu') >= 0);
                for (const el of roots) {
                    let menu = null;
                    try { menu = $find(el.id); } catch (e) {}
                    if (!menu || typeof menu.get_items !== 'function') continue;
                    result.menus++;
                    for (const it of walk(menu.get_items(), [])) {
                        const text = typeof it.get_text === 'function' ? it.get_text() : '';
                        if (text) result.options.push(text.trim());
                        if (norm(text) !== target) continue;
                        result.found = true;
                        try {
                            const parent = typeof it.get_parent === 'function' ? it.get_parent() : null;
                            if (parent && typeof parent.open === 'function') parent.open();
                        } catch (e) {}
                        try {
                            it.click();
                            result.method = 'RadMenuItem.click';
                            return result;
                        } catch (e) { result.error = String(e); }
                    }
                }
            }

            // Telerik's delegated handler wants the mouseover/mousedown/mouseup
            // that precede a real click - el.click() alone sends none of them.
            const hit = Array.from(document.querySelectorAll('.rmLink .rmText'))
                .find(el => norm(el.textContent) === target);
            if (!hit) return result;
            result.found = true;
            const link = hit.closest('a.rmLink') || hit;
            for (const type of ['mouseover', 'mousedown', 'mouseup', 'click']) {
                link.dispatchEvent(new MouseEvent(type, {bubbles: true, cancelable: true, view: window}));
            }
            result.method = 'mouse events';
            return result;
        }""",
        EXPECTED_CENTER,
    )

    if not switch.get("found"):
        raise Exception(
            f"Center scope is '{current}', expected '{EXPECTED_CENTER}', and no menu item "
            f"matches it. Refusing to export - the CSV would be scoped to the wrong "
            f"center. Menu items seen: {switch.get('options')[:40]}"
        )
    print(f"  Center switch dispatched via {switch.get('method')} "
          f"(RadMenu objects found: {switch.get('menus')})")

    settle(page)
    time.sleep(3)

    frame, sel = find_in_frames(page, root_sels, timeout=60000)
    if not frame:
        raise Exception("Center picker disappeared after the switch; cannot confirm the center scope")
    after = (frame.locator(sel).first.text_content() or "").strip()
    if _norm_center(after) != _norm_center(EXPECTED_CENTER):
        raise Exception(
            f"Center switch did not take: picker still reads '{after}', expected "
            f"'{EXPECTED_CENTER}'. Dispatched via {switch.get('method')}, "
            f"RadMenu objects found: {switch.get('menus')}, error: {switch.get('error')}"
        )
    print(f"  Center scope now: '{after}'")


def select_custom_time_period(page, frame):
    print("Setting Time Period = Custom...")
    frame.locator("#ddlTimePeriod").click()
    time.sleep(1.5)

    # The option list is a zfc dropdown rendered outside #ddlTimePeriod, so match
    # on the visible label text instead of a container-scoped selector.
    option = frame.locator(".divFilterListItm .zfcListItmLbl").filter(has_text="Custom")
    clicked = False
    try:
        option.first.wait_for(state="visible", timeout=15000)
        option.first.click(timeout=10000)
        clicked = True
    except Exception as e:
        print(f"  Label click failed ({e}); dispatching click from JS...")
        clicked = frame.evaluate("""() => {
            const lbl = Array.from(document.querySelectorAll('.divFilterListItm .zfcListItmLbl'))
                .find(el => el.textContent.trim() === 'Custom');
            if (!lbl) return false;
            const item = lbl.closest('.divFilterListItm') || lbl;
            item.click();
            return true;
        }""")

    if not clicked:
        raise Exception("'Custom' option not found in the Time Period dropdown")

    time.sleep(2)
    # Close the dropdown so the From/To inputs are not covered.
    page.keyboard.press("Escape")
    page.mouse.click(5, 5)
    time.sleep(1)

    frame.wait_for_selector("#dateContainer", state="visible", timeout=30000)
    print("  Time Period = Custom applied.")


def set_date_input(frame, selector, value):
    el = frame.locator(selector)
    el.wait_for(state="visible", timeout=30000)
    el.click()
    el.fill("")
    el.press_sequentially(value, delay=60)
    time.sleep(0.5)
    el.press("Escape")

    # jQuery UI datepicker inputs only commit on change/blur; force both, and
    # push the value through the datepicker API so its internal state matches.
    frame.evaluate(
        """([sel, val]) => {
            const el = document.querySelector(sel);
            if (!el) return null;
            el.value = val;
            if (window.jQuery) {
                const $el = window.jQuery(el);
                try { $el.datepicker('setDate', val); } catch (e) {}
                $el.trigger('change');
                $el.trigger('blur');
            } else {
                el.dispatchEvent(new Event('input', {bubbles: true}));
                el.dispatchEvent(new Event('change', {bubbles: true}));
                el.dispatchEvent(new Event('blur', {bubbles: true}));
            }
            return el.value;
        }""",
        [selector, value],
    )
    time.sleep(1.5)

    actual = el.input_value()
    if actual.strip() != value:
        raise Exception(f"Date not applied on {selector}: asked {value}, field holds '{actual}'")
    print(f"  {selector} = {actual}")


def set_dates(frame, from_text, to_text):
    print(f"Setting date range: {from_text} to {to_text}")
    set_date_input(frame, "#MainContent_MainContent_PageContent_dpFromDate", from_text)
    set_date_input(frame, "#MainContent_MainContent_PageContent_dpToDate", to_text)

    # Some listing pages reload the grid only on an explicit refresh/apply.
    frame.evaluate("""() => {
        const btn = document.querySelector('#btnRefresh, #MainContent_MainContent_PageContent_btnRefresh, #btnApply');
        if (btn) { btn.removeAttribute('disabled'); btn.click(); return true; }
        return false;
    }""")
    time.sleep(2)


def wait_for_grid(page, frame, timeout=300000):
    """Rendered row count, 0 for a confirmed-empty grid, -1 if it never settled."""
    print("Waiting for the table to load...")
    settle(page, timeout=timeout)

    deadline = time.time() + timeout / 1000.0
    empty_streak = 0
    last = None
    while time.time() < deadline:
        state = frame.evaluate("""() => {
            const q = s => document.querySelector(s);

            // offsetParent alone is not enough: ag-Grid hides overlays with the
            // ag-hidden class and with visibility/opacity on ancestors.
            const isVisible = el => {
                if (!el) return false;
                const r = el.getBoundingClientRect();
                if (r.width === 0 && r.height === 0) return false;
                for (let n = el; n; n = n.parentElement) {
                    if (n.classList && n.classList.contains('ag-hidden')) return false;
                    const st = getComputedStyle(n);
                    if (st.display === 'none' || st.visibility === 'hidden') return false;
                    if (parseFloat(st.opacity) === 0) return false;
                }
                return true;
            };

            const rows = document.querySelectorAll(
                '.ag-center-cols-container .ag-row, .ag-body-viewport .ag-row, ' +
                'table.gridTable tbody tr, #gridPOList .ag-row'
            ).length;
            const center = q('.ag-center-cols-container');

            const loadingEl = q('.ag-overlay-loading-center, .ag-overlay-loading-wrapper');
            // This build renders no-rows as .ag-overlay-no-rows-wrapper wrapping
            // .ag-overlay-nodata ("No records to display"); older ones use
            // .ag-overlay-no-rows-center. Accept any of them, plus the text.
            const emptyEl = q(
                '.ag-overlay-nodata, .ag-overlay-no-rows-wrapper, .ag-overlay-no-rows-center'
            );
            const emptyVisible = isVisible(emptyEl);
            const overlayText = emptyVisible ? (emptyEl.textContent || '').trim().slice(0, 80) : '';

            return {
                rows,
                // Grid shell present at all? Distinguishes "no data" from "not rendered yet".
                mounted: !!q('.ag-body-viewport') && !!center,
                // An empty result also collapses every row container to height: 1px.
                centerHeight: center ? Math.round(center.getBoundingClientRect().height) : null,
                loading: isVisible(loadingEl),
                empty: emptyVisible,
                overlayText,
            };
        }""")
        last = state

        if state["loading"]:
            empty_streak = 0
            time.sleep(2)
            continue

        if state["rows"] > 0:
            print(f"  Table loaded: {state['rows']} rendered row(s).")
            return state["rows"]

        collapsed = state["centerHeight"] is not None and state["centerHeight"] <= 2
        if state["mounted"] and (state["empty"] or collapsed):
            # Require repeat readings: mid-refresh the grid also looks empty.
            empty_streak += 1
            if empty_streak >= GRID_EMPTY_CONFIRMATIONS:
                reason = (
                    f"overlay says '{state['overlayText']}'"
                    if state["empty"]
                    else f"row container collapsed to {state['centerHeight']}px"
                )
                print(f"  Table loaded with no rows ({reason}).")
                return 0
        else:
            empty_streak = 0

        time.sleep(2)

    if last and last["mounted"] and last["rows"] == 0:
        print(f"  Grid still empty at timeout (last: {last}); treating as no data.")
        return 0

    print(f"  WARNING: grid state never settled (last: {last}); exporting anyway.")
    return -1


def export_orders_csv(page, frame, end_date):
    print("Exporting to CSV...")
    frame.locator("#gridExportBtn").click()
    time.sleep(1.5)
    frame.wait_for_selector("#MainContent_MainContent_PageContent_btnExportcsv", state="attached", timeout=30000)

    # The export link is a __doPostBack href; click it from JS so an overlay on
    # the dropdown cannot swallow the event.
    with page.expect_download(timeout=600000) as download_info:
        frame.evaluate("""() => {
            const a = document.querySelector('#MainContent_MainContent_PageContent_btnExportcsv');
            if (a) { a.click(); return; }
            if (typeof __doPostBack === 'function') {
                __doPostBack('ctl00$ctl00$ctl00$MainContent$MainContent$PageContent$btnExportcsv', '');
            }
        }""")

    download = download_info.value
    script_dir = os.path.dirname(__file__) or "."
    # Always .csv: the link clicked is btnExportcsv, and
    # validate_report_folders() has to match one exact name.
    filename = os.path.join(script_dir, f"orders_{end_date}.csv")
    download.save_as(filename)
    time.sleep(2)
    print(f"  Saved as: {filename} (server suggested: {download.suggested_filename})")
    return filename


def download_orders(page, start_date, end_date):
    # No center check here: main() enforces the scope once after login, before
    # any report runs.
    open_orders_page(page)
    frame, _ = find_in_frames(page, ("#ddlTimePeriod",), timeout=120000)
    if not frame:
        raise Exception(f"#ddlTimePeriod missing on the Orders page. URL: {page.url}")
    select_custom_time_period(page, frame)
    set_dates(frame, _mdy(start_date), _mdy(end_date))
    rows = wait_for_grid(page, frame)

    if rows == 0:
        # Nothing to export: the CSV button on an empty grid yields a
        # headers-only (or error) file, which is worse than no file at all.
        print(f"No orders for {_mdy(start_date)} to {_mdy(end_date)}. Skipping the export.")
        return ORDERS_NO_DATA

    filename = export_orders_csv(page, frame, end_date)

    print(f"Validating downloaded file: {filename}")
    validate_csv(filename)
    print(f"Downloaded: {filename}")
    return filename


def report_window(report_name):
    """(start, end) date strings for a report's date filter."""
    if report_name == "Business KPI":
        return BKP_START_DATE, END_DATE
    if report_name == "Orders":
        return ORDERS_START_DATE, END_DATE  # rolling calendar month ending T-1
    if report_name == "Current Stock":
        return END_DATE, END_DATE           # single day: T-1
    return START_DATE, END_DATE


REPORT_FILTERS = {
    "Appointments": apply_appointments_filters,
    "Attendance": apply_attendance_filters,
    "Cost of Goods": apply_cost_of_goods_filters,
    "Sales-Accrual": apply_sales_accrual_filters,
    "Employee Sales": apply_employee_sales_filters,
    "Sales-Cash": apply_sales_cash_filters,
    "Business KPI": apply_business_kpi_filters,
    "Memberships": apply_memberships_filters,
    "Current Stock": apply_current_stock_filters,
    "Stock Ledger": apply_stock_ledger_filters,
}


def download_report(context, page, report_name, start_date, end_date):
    if report_name == "Orders":
        # Not a ReportsDashboard report: own nav, own filters, own export link.
        # Returns ORDERS_NO_DATA when the window holds no orders.
        return download_orders(page, start_date, end_date)

    page.goto("https://evolvemedspa.zenoti.com/Admin/Reports/ReportsDashboard.aspx")
    settle(page, timeout=45000)
    time.sleep(2)
    print(f"Opening report: {report_name}")

    # Reports reachable from the bookmarks grid: Zenoti's own row onclick is
    # ReportsGrid_Row_Click(event,'<id>'), so call it directly. Dispatching a
    # click on the row instead goes through the #dialog-reports modal overlay,
    # which is how Employee Sales ended up printing its onclick attribute and
    # then never opening a tab.
    BOOKMARKED_REPORT_IDS = {
        "Business KPI": "business_kpi",
        "Memberships": "memberships",
        "Stock Ledger": "stock_ledger",
        "Current Stock": "current_stock",
        "Employee Sales": "employee_sales",
    }

    report_id = BOOKMARKED_REPORT_IDS.get(report_name)
    if report_id:
        page.evaluate('loadBookmarksViewAllGrid("Bookmarked")')
        time.sleep(3)
        with context.expect_page(timeout=120000) as new_page_info:
            page.evaluate(f"ReportsGrid_Row_Click(event,'{report_id}')")
    else:
        with context.expect_page(timeout=120000) as new_page_info:
            page.locator('#gridReports span.report-name').get_by_text(report_name, exact=True).click(timeout=60000)

    time.sleep(2)
    report_page = new_page_info.value
    # Non-fatal: Zenoti report pages hold long-poll connections open, so
    # networkidle can legitimately never arrive - on Railway's slower link that
    # turned into a bare "Timeout 120000ms exceeded" that failed the report.
    settle(report_page, timeout=45000)
    time.sleep(2)
    print(f"{report_name} report page loaded.")

    if report_name == "Sales-Accrual":
        start_dt = f"{start_date} 00:00"
        end_dt = f"{end_date} 23:59"
        dt_format = "YYYY-MM-DD HH:mm"
    elif report_name == "Current Stock":
        # Single "Stock as on" datetime = T-1 at 07:00 AM
        start_dt = f"{start_date} 07:00"
        end_dt = f"{start_date} 07:00"
        dt_format = "YYYY-MM-DD HH:mm"
    else:
        start_dt = start_date
        end_dt = end_date
        dt_format = "YYYY-MM-DD"

    report_page.evaluate(f"""
        (function() {{
            var picker = $('#elm_dates').data('daterangepicker');
            if (picker) {{
                var startDate = moment('{start_dt}', '{dt_format}');
                var endDate = moment('{end_dt}', '{dt_format}');
                picker.setStartDate(startDate);
                picker.setEndDate(endDate);
                picker.element.trigger('apply.daterangepicker', picker);
            }}
        }})();
    """)
    time.sleep(2)
    print("Date range set.")

    filter_fn = REPORT_FILTERS.get(report_name)
    if filter_fn:
        filter_fn(report_page)
    time.sleep(2)

    print("Refreshing report...")
    if report_name == "Current Stock":
        report_page.evaluate("""
            (function() {
                if (typeof btnCurrentStock_onClick === 'function') {
                    btnCurrentStock_onClick();
                } else {
                    var b = document.querySelector('#btnRefresh');
                    if (b) b.click();
                }
            })();
        """)
    elif report_name == "Employee Sales":
        # #btnRefresh renders disabled until the page decides the filter set is
        # complete; drop the attribute so the click registers.
        report_page.evaluate("""
            (function() {
                var b = document.querySelector('#btnRefresh');
                if (b) {
                    b.removeAttribute('disabled');
                    b.click();
                }
            })();
        """)
    else:
        report_page.evaluate("document.querySelector('#btnRefresh').click()")
    time.sleep(2)
    report_page.wait_for_load_state("networkidle", timeout=300000)
    time.sleep(2)

    # The dropdown offers CSV (#export_csv), Excel (#export_excel_v2) and a hidden
    # "Excel with subtotals"; #export_excel_v2 runs Export_Click(true, "excel_v2").
    # Current Stock's Excel export is gated behind CURRENT_STOCK_AS_EXCEL.
    is_excel = report_name == "Current Stock" and CURRENT_STOCK_AS_EXCEL
    export_sel = "#export_excel_v2" if is_excel else "#export_csv"

    print(f"Exporting report to {'Excel' if is_excel else 'CSV'}...")
    report_page.locator('#dropdownMenuLink').click()
    time.sleep(2)
    report_page.wait_for_selector(export_sel, state='attached', timeout=30000)

    download_timeout = 900000 if report_name == "Stock Ledger" else 600000 if report_name == "Employee Sales" else 300000
    with report_page.expect_download(timeout=download_timeout) as download_info:
        report_page.evaluate(f"document.querySelector('{export_sel}').click()")

    time.sleep(10)
    download = download_info.value
    script_dir = os.path.dirname(__file__) or "."
    if report_name == "Business KPI":
        filename = os.path.join(script_dir, f"business_kpi_{start_date}_to_{end_date}.csv")
    elif report_name == "Current Stock":
        if is_excel:
            # Trust the server's extension when it gives one (.xlsx vs .xls);
            # validate_report_folders() accepts either.
            suggested = download.suggested_filename or ""
            ext = os.path.splitext(suggested)[1].lower()
            if ext not in (".xlsx", ".xls"):
                ext = ".xlsx"
        else:
            ext = ".csv"
        filename = os.path.join(script_dir, f"current_stock_{end_date}{ext}")
    else:
        safe_name = report_name.replace(" ", "_").lower()
        filename = os.path.join(script_dir, f"{safe_name}_{start_date}_to_{end_date}.csv")
    download.save_as(filename)
    time.sleep(2)

    print(f"Validating downloaded file: {filename}")
    if is_excel:
        validate_excel(filename)
    else:
        validate_csv(filename)
    print(f"Downloaded: {filename}")

    report_page.close()
    time.sleep(2)
    page.bring_to_front()
    time.sleep(2)
    return filename


print("Script starting...")
sys.stdout.flush()

LOG_FILENAME = os.path.join(os.path.dirname(__file__) or ".", f"download_report_logs_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.txt")
log_file = open(LOG_FILENAME, "w", encoding="utf-8")


class Tee:
    def __init__(self, *streams):
        self.streams = streams

    def write(self, data):
        for s in self.streams:
            s.write(data)
            s.flush()

    def flush(self):
        for s in self.streams:
            s.flush()


sys.stdout = Tee(sys.__stdout__, log_file)

cleanup_old_csvs()

# Phase 1: scan every report folder upfront and move any existing file
# (1 or more) to Done before downloading anything. Runs before the browser
# launches so leftovers from a previous run are cleared even if login fails,
# and so a bad GOOGLE_TOKEN_JSON surfaces before a full login.
print("Moving existing report files to Done...")
move_existing_reports_to_done()

with sync_playwright() as p:
    print("Playwright started.")
    browser, context = create_browser_and_context(p)
    print("Browser launched.")
    page = context.new_page()

    try:
        if needs_login(page):
            print("No valid session. Logging in...")
            do_login(page)
            save_cookies(context)
        else:
            print("Session valid from saved cookies. Skipping login.")

        wait_for_dashboard(page)
        save_cookies(context)

        # Every report's export inherits the RadMenu center scope, so set it once
        # up front rather than only before Orders. A single center selected there
        # silently scopes every CSV to that center under a whole-org filename.
        ensure_center_scope(page)
        save_cookies(context)

        reports = ["Stock Ledger", "Appointments", "Sales-Cash", "Cost of Goods", "Attendance", "Business KPI", "Memberships", "Current Stock", "Employee Sales", "Orders"]
        # reports = ["Employee Sales"]
        failed_reports = []
        succeeded_reports = []

        for report in reports:
            try:
                report_start, report_end = report_window(report)
                if report == "Business KPI":
                    print(f"Business KPI date filter: {report_start} to {report_end}")
                filename = download_report(context, page, report, report_start, report_end)
                if filename is ORDERS_NO_DATA:
                    # Not a failure and not a success: nothing was uploaded, so
                    # keeping it out of succeeded_reports keeps
                    # validate_report_folders() from hunting for a file that was
                    # never meant to exist.
                    print(f"{report}: no rows for {report_start}; nothing to upload.")
                    save_cookies(context)
                    continue
                folder_id = REPORT_FOLDERS.get(report, DRIVE_FOLDER_ID)
                upload_to_drive(filename, folder_id)
                os.remove(filename)
                save_cookies(context)
                succeeded_reports.append(report)
                time.sleep(5)
            except Exception as e:
                print(f"FAILED: {report} — {e}")
                failed_reports.append((report, str(e)))
                for p in context.pages:
                    if p != page:
                        try:
                            p.close()
                        except Exception:
                            pass
                page.bring_to_front()
                time.sleep(2)

        # Bounded retry passes. One pass was not enough: the failures seen so far
        # are transient timeouts that a fresh nav clears.
        for attempt in range(1, RETRY_MAX_ATTEMPTS + 1):
            if not failed_reports:
                break

            print(f"\n--- Retrying {len(failed_reports)} failed report(s) "
                  f"(attempt {attempt}/{RETRY_MAX_ATTEMPTS}) ---")
            relogin_ok = True
            try:
                if needs_login(page):
                    print("Re-logging in before retry...")
                    do_login(page)
                    save_cookies(context)
                    wait_for_dashboard(page)
                    # Fresh session, fresh center scope.
                    ensure_center_scope(page)
            except Exception as e:
                print(f"Re-login failed, skipping retries: {e}")
                relogin_ok = False

            retry_still_failed = []
            if not relogin_ok:
                retry_still_failed = list(failed_reports)
            for report, prev_error in (failed_reports if relogin_ok else []):
                try:
                    print(f"Retrying: {report}")
                    report_start, report_end = report_window(report)
                    filename = download_report(context, page, report, report_start, report_end)
                    if filename is ORDERS_NO_DATA:
                        print(f"{report}: no rows for {report_start}; nothing to upload.")
                        save_cookies(context)
                        continue
                    folder_id = REPORT_FOLDERS.get(report, DRIVE_FOLDER_ID)
                    upload_to_drive(filename, folder_id)
                    os.remove(filename)
                    save_cookies(context)
                    succeeded_reports.append(report)
                    time.sleep(5)
                except Exception as e:
                    print(f"RETRY FAILED: {report} — {e}")
                    retry_still_failed.append((report, str(e)))
                    for p in context.pages:
                        if p != page:
                            try:
                                p.close()
                            except Exception:
                                pass
                    page.bring_to_front()
                    time.sleep(2)

            failed_reports = retry_still_failed
            if not relogin_ok:
                break
            if failed_reports and attempt < RETRY_MAX_ATTEMPTS:
                print(f"Still failing {[r for r, _ in failed_reports]}; "
                      f"next pass in {REDOWNLOAD_RETRY_BACKOFF}s...")
                time.sleep(REDOWNLOAD_RETRY_BACKOFF)

        print(f"\n--- Report Summary ---")
        print(f"Succeeded: {succeeded_reports}")
        if failed_reports:
            print(f"Failed: {[r for r, _ in failed_reports]}")

        print(f"\n--- Google Sheet Extraction ---")
        gsheet_succeeded, gsheet_failed = download_gsheet_reports()
        print(f"Google Sheets succeeded: {gsheet_succeeded}")
        if gsheet_failed:
            print(f"Google Sheets failed: {[r for r, _ in gsheet_failed]}")

        print("Checking report folders for duplicate filenames...")
        dedupe_report_folders()

        print("Validating report folders contain recent downloads...")
        missing_reports = validate_report_folders(succeeded_reports)

        if missing_reports:
            print(f"\n--- Re-checking {len(missing_reports)} missing report(s) before re-download ---")
            time.sleep(10)
            missing_reports = validate_report_folders(missing_reports)

        # Re-download anything Drive says is absent, then dedupe + re-validate and
        # try again. Bounded at REDOWNLOAD_MAX_ATTEMPTS: a report that fails twice
        # is failing for a reason a third pass won't fix, and each pass costs a
        # full download.
        redownload_failed = []
        attempted_redownload = bool(missing_reports)
        for attempt in range(1, REDOWNLOAD_MAX_ATTEMPTS + 1):
            if not missing_reports:
                break

            print(
                f"\n--- Re-downloading {len(missing_reports)} missing report(s) "
                f"(attempt {attempt}/{REDOWNLOAD_MAX_ATTEMPTS}) ---"
            )
            try:
                if needs_login(page):
                    print("Re-logging in before re-download...")
                    do_login(page)
                    save_cookies(context)
                    wait_for_dashboard(page)
                    # Fresh session, fresh center scope.
                    ensure_center_scope(page)
            except Exception as e:
                print(f"Re-login failed, abandoning re-download: {e}")
                break

            redownload_failed = []
            for report in list(missing_reports):
                try:
                    print(f"Re-downloading: {report}")
                    report_start, report_end = report_window(report)
                    filename = download_report(context, page, report, report_start, report_end)
                    if filename is ORDERS_NO_DATA:
                        print(f"{report}: no rows for {report_start}; nothing to upload.")
                        save_cookies(context)
                        continue
                    folder_id = REPORT_FOLDERS.get(report, DRIVE_FOLDER_ID)
                    upload_to_drive(filename, folder_id)
                    os.remove(filename)
                    save_cookies(context)
                    time.sleep(2)
                except Exception as e:
                    print(f"RE-DOWNLOAD FAILED: {report} — {e}")
                    redownload_failed.append(report)
                    for p in context.pages:
                        if p != page:
                            try:
                                p.close()
                            except Exception:
                                pass
                    page.bring_to_front()
                    time.sleep(2)

            if redownload_failed:
                print(f"Re-download failures: {redownload_failed}")

            # Validate against the reports we just re-downloaded, not the full
            # succeeded list, so the next pass retries only what is still absent.
            print(f"Validating after re-download attempt {attempt}...")
            dedupe_report_folders()
            missing_reports = validate_report_folders(missing_reports)

            if missing_reports and attempt < REDOWNLOAD_MAX_ATTEMPTS:
                print(f"Still missing {missing_reports}; retrying in {REDOWNLOAD_RETRY_BACKOFF}s...")
                time.sleep(REDOWNLOAD_RETRY_BACKOFF)

        # Final sweep across every report that downloaded successfully, so one that
        # vanished from Drive after its own pass is still caught. Skipped when the
        # first validation was already clean — line above would just repeat it.
        if attempted_redownload:
            print("Final validation...")
            dedupe_report_folders()
            still_missing = validate_report_folders(succeeded_reports)
        else:
            still_missing = missing_reports

        if still_missing:
            raise Exception(
                f"Reports still missing after {REDOWNLOAD_MAX_ATTEMPTS} re-download attempt(s): {still_missing}"
            )

        print("Logging out...")
        page.goto("https://evolvemedspa.zenoti.com/Admin/Reports/ReportsDashboard.aspx")
        page.wait_for_load_state("networkidle", timeout=60000)
        time.sleep(1)
        page.locator('#usernameBtn').click
        time.sleep(1)
        page.locator('.userLogoutCls').click
        time.sleep(2)
        print("Logged out.")

        if failed_reports:
            raise Exception(f"Reports failed after retry: {[r for r, _ in failed_reports]}")

    except Exception as e:
        print(f"Error: {e}")
        raise
    finally:
        context.close()
        browser.close()

    sys.stdout = sys.__stdout__
    log_file.close()
    upload_to_drive(LOG_FILENAME, DRIVE_FOLDER_ID)
    os.remove(LOG_FILENAME)